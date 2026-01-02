"""
Excel export functionality for evaluation results.
"""

import os
import logging
from datetime import datetime
from typing import List, Dict

from ..booking import CONVERSATION_STEPS, get_stage_progress, get_failed_at_description
from .excel_styles import (
    apply_header_style,
    apply_rate_color,
    apply_result_color,
    apply_stage_color,
)
from .charts import create_trend_chart
from .visualization import create_run_visualization_sheet

logger = logging.getLogger("eval-runner")

DEFAULT_RESULTS_FILE = "evaluation_results.xlsx"

# Column widths for sheets
SUMMARY_WIDTHS = {"A": 8, "B": 20, "C": 12, "D": 12, "E": 10, "F": 10, "G": 18}
DETAIL_WIDTHS = [8, 20, 30, 40, 30, 12, 10, 20, 15, 18, 15, 14, 14, 14, 50, 10]


def update_results_excel(
    results: List[Dict],
    excel_file: str = DEFAULT_RESULTS_FILE,
) -> str:
    """
    Update the persistent Excel file with new evaluation results.
    Creates the file if it doesn't exist, appends if it does.
    Includes a trend chart showing success rate over time.

    Args:
        results: List of scenario results
        excel_file: Path to the Excel file

    Returns:
        The path to the Excel file, or empty string on failure
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed. Run: pip install openpyxl")
        logger.warning("Skipping Excel export.")
        return ""

    # Calculate summary for this run
    run_timestamp = datetime.now()
    total = len(results)
    passed = sum(1 for r in results if r.get("success_results", {}).get("booking_confirmed", False))
    failed = total - passed
    success_rate = (passed / total * 100) if total > 0 else 0

    # Prepare run summary
    run_summary = {
        "timestamp": run_timestamp.strftime("%Y-%m-%d %H:%M:%S"),
        "date": run_timestamp.strftime("%Y-%m-%d"),
        "total_scenarios": total,
        "passed": passed,
        "failed": failed,
        "success_rate": round(success_rate, 1),
    }

    # Check if file exists
    file_exists = os.path.exists(excel_file)

    if file_exists:
        wb = openpyxl.load_workbook(excel_file)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Update each sheet
    run_number, next_row = _update_summary_sheet(wb, run_summary)
    _update_details_sheet(wb, results, run_number, run_summary)
    _update_stages_sheet(wb, results, run_number, run_summary)
    create_run_visualization_sheet(wb, results, run_number, run_summary)

    # Save file
    wb.save(excel_file)
    logger.info(f"Excel results updated: {excel_file}")
    logger.info(f"   Run #{run_number}: {passed}/{total} passed ({success_rate:.1f}%)")

    return excel_file


def _update_summary_sheet(wb, run_summary: Dict) -> tuple:
    """
    Update the Runs Summary sheet.

    Args:
        wb: openpyxl Workbook
        run_summary: Summary dict for this run

    Returns:
        Tuple of (run_number, next_row)
    """
    from openpyxl.styles import Alignment

    if "Runs Summary" not in wb.sheetnames:
        ws = wb.create_sheet("Runs Summary", 0)
        headers = ["Run #", "Timestamp", "Date", "Scenarios", "Passed", "Failed", "Success Rate (%)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            apply_header_style(cell)

        # Set column widths
        for col_letter, width in SUMMARY_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

        next_row = 2
        run_number = 1
    else:
        ws = wb["Runs Summary"]
        next_row = ws.max_row + 1
        run_number = next_row - 1

    # Add this run's summary
    ws.cell(row=next_row, column=1, value=run_number)
    ws.cell(row=next_row, column=2, value=run_summary["timestamp"])
    ws.cell(row=next_row, column=3, value=run_summary["date"])
    ws.cell(row=next_row, column=4, value=run_summary["total_scenarios"])
    ws.cell(row=next_row, column=5, value=run_summary["passed"])
    ws.cell(row=next_row, column=6, value=run_summary["failed"])
    ws.cell(row=next_row, column=7, value=run_summary["success_rate"])

    # Center align data cells
    for col in range(1, 8):
        ws.cell(row=next_row, column=col).alignment = Alignment(horizontal="center")

    # Color code success rate
    rate_cell = ws.cell(row=next_row, column=7)
    apply_rate_color(rate_cell, run_summary["success_rate"])

    # Update trend chart
    create_trend_chart(ws, next_row)

    return run_number, next_row


def _update_details_sheet(wb, results: List[Dict], run_number: int, run_summary: Dict):
    """
    Update the All Results sheet with detailed scenario results.

    Args:
        wb: openpyxl Workbook
        results: List of scenario results
        run_number: Current run number
        run_summary: Summary dict for this run
    """
    from openpyxl.utils import get_column_letter

    # Collect all unique evaluation criteria from all scenarios
    all_criteria = set()
    for r in results:
        scenario = r.get("scenario", {})
        criteria = scenario.get("evaluation_criteria", {})
        all_criteria.update(criteria.keys())

    # Sort criteria for consistent column order
    sorted_criteria = sorted(all_criteria)

    if "All Results" not in wb.sheetnames:
        ws = wb.create_sheet("All Results")
        headers = [
            "Run #", "Timestamp", "Scenario ID", "Scenario Name", "Customer",
            "Duration (s)", "Messages", "Final Stage", "Stage Progress",
            "Booking Confirmed", "Booking Number",
            "Name Provided", "Phone Provided", "Email Provided",
        ]

        # Add evaluation criteria headers
        for criterion in sorted_criteria:
            # Format criterion name for header (replace underscores with spaces, title case)
            header_name = criterion.replace("_", " ").title()
            headers.append(header_name)

        headers.extend(["Error Description", "Result"])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            apply_header_style(cell)

        # Set column widths - use 15 for criteria columns
        base_widths = [8, 20, 30, 40, 30, 12, 10, 20, 15, 18, 15, 14, 14, 14]
        criteria_widths = [15] * len(sorted_criteria)
        final_widths = [50, 10]  # Error Description, Result
        all_widths = base_widths + criteria_widths + final_widths

        for i, width in enumerate(all_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        detail_row = 2
    else:
        ws = wb["All Results"]
        detail_row = ws.max_row + 1

    # Add each result
    for r in results:
        success = r.get("success_results", {})
        is_passed = success.get("booking_confirmed", False)
        stage = success.get("conversation_stage", "UNKNOWN")
        provided_info = success.get("provided_info", {})

        # Calculate stage progress
        step_num, total_steps = get_stage_progress(stage)
        progress_str = f"{step_num}/{total_steps}"

        # Get error/warning description
        error = r.get("error", "")
        if not error:
            if is_passed and success.get("invalid_booking_number"):
                # Booking succeeded but number format was invalid - show as warning
                invalid_value = success.get("invalid_booking_number_value", "unknown")
                error = f"Warning: Booking number format invalid: '{invalid_value}'"
            elif not is_passed:
                # Booking failed - get the failure reason
                error = get_failed_at_description(stage, r.get("transcripts", []))

        # Get customer name
        customer_name = ""
        scenario_name = r.get("scenario_name", "")
        if "-" in scenario_name:
            customer_name = scenario_name.split("-")[0]
        else:
            # Try to get from scenario customer data
            scenario = r.get("scenario", {})
            customer = scenario.get("customer", {})
            customer_name = customer.get("name", "")

        # Write base columns
        ws.cell(row=detail_row, column=1, value=run_number)
        ws.cell(row=detail_row, column=2, value=run_summary["timestamp"])
        ws.cell(row=detail_row, column=3, value=r.get("scenario_id", ""))
        ws.cell(row=detail_row, column=4, value=scenario_name)
        ws.cell(row=detail_row, column=5, value=customer_name)
        ws.cell(row=detail_row, column=6, value=r.get("duration_seconds", 0))
        ws.cell(row=detail_row, column=7, value=r.get("transcript_count", 0))
        ws.cell(row=detail_row, column=8, value=stage)
        ws.cell(row=detail_row, column=9, value=progress_str)

        # Booking Confirmed with color
        booking_cell = ws.cell(row=detail_row, column=10, value="YES" if is_passed else "NO")
        apply_result_color(booking_cell, is_passed)

        # Show the booking number - use whatever the agent said (valid or invalid)
        booking_num = success.get("booking_number", "") or success.get("raw_booking_number", "") or ""
        ws.cell(row=detail_row, column=11, value=booking_num)
        ws.cell(row=detail_row, column=12, value="YES" if provided_info.get("name") else "NO")
        ws.cell(row=detail_row, column=13, value="YES" if provided_info.get("phone") else "NO")
        ws.cell(row=detail_row, column=14, value="YES" if provided_info.get("email") else "NO")

        # Add evaluation criteria results
        col_offset = 15
        scenario = r.get("scenario", {})
        scenario_criteria = scenario.get("evaluation_criteria", {})
        criteria_results = r.get("criteria_results", {})

        for criterion in sorted_criteria:
            # Check if this criterion applies to this scenario
            if criterion in scenario_criteria:
                # Get the result for this criterion (default to "N/A" if not evaluated yet)
                criterion_result = criteria_results.get(criterion, "N/A")
                criterion_cell = ws.cell(row=detail_row, column=col_offset, value=criterion_result)

                # Color code: PASS = green, FAIL = red, N/A = gray
                if criterion_result == "PASS":
                    apply_result_color(criterion_cell, True)
                elif criterion_result == "FAIL":
                    apply_result_color(criterion_cell, False)
            else:
                # This criterion doesn't apply to this scenario
                ws.cell(row=detail_row, column=col_offset, value="-")

            col_offset += 1

        # Error Description and Result columns
        ws.cell(row=detail_row, column=col_offset, value=error or "")
        result_cell = ws.cell(row=detail_row, column=col_offset + 1, value="PASS" if is_passed else "FAIL")
        apply_result_color(result_cell, is_passed)

        # Color code stage
        stage_cell = ws.cell(row=detail_row, column=8)
        apply_stage_color(stage_cell, stage, step_num)

        detail_row += 1


def _update_stages_sheet(wb, results: List[Dict], run_number: int, run_summary: Dict):
    """
    Update the Stage Breakdown sheet.

    Args:
        wb: openpyxl Workbook
        results: List of scenario results
        run_number: Current run number
        run_summary: Summary dict for this run
    """
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from .excel_styles import get_success_fill, get_failure_fill

    if "Stage Breakdown" not in wb.sheetnames:
        ws = wb.create_sheet("Stage Breakdown")
        headers = ["Run #", "Timestamp"] + CONVERSATION_STEPS + ["Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            apply_header_style(cell)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Set column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 20
        for i in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 12
        ws.row_dimensions[1].height = 40

        stage_row = 2
    else:
        ws = wb["Stage Breakdown"]
        stage_row = ws.max_row + 1

    # Count scenarios at each stage
    stage_counts = {stage: 0 for stage in CONVERSATION_STEPS}
    for r in results:
        stage = r.get("success_results", {}).get("conversation_stage", "GREETING")
        if stage in stage_counts:
            stage_counts[stage] += 1

    # Add row
    ws.cell(row=stage_row, column=1, value=run_number)
    ws.cell(row=stage_row, column=2, value=run_summary["timestamp"])

    for i, stage in enumerate(CONVERSATION_STEPS, 3):
        count = stage_counts.get(stage, 0)
        cell = ws.cell(row=stage_row, column=i, value=count)
        cell.alignment = Alignment(horizontal="center")

        # Color code
        if stage == "BOOKING_CONFIRMED" and count > 0:
            cell.fill = get_success_fill()
        elif count > 0:
            cell.fill = get_failure_fill()

    ws.cell(row=stage_row, column=len(CONVERSATION_STEPS) + 3, value=run_summary["total_scenarios"])
