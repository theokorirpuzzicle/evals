"""
Visual dashboard creation for evaluation runs.
Creates colorful heatmaps, gauges, and charts for easy interpretation.
"""

import logging
from typing import List, Dict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("visualization")

# Color scheme
COLOR_PASS = "00C853"  # Green
COLOR_FAIL = "D32F2F"  # Red
COLOR_NA = "9E9E9E"    # Gray
COLOR_HEADER = "1976D2"  # Blue
COLOR_SUBHEADER = "64B5F6"  # Light Blue
COLOR_LEGEND_BG = "E3F2FD"  # Very Light Blue


def create_run_visualization_sheet(wb, results: List[Dict], run_number: int, run_summary: Dict):
    """
    Create or update the Run Visualization sheet with a colorful dashboard.

    Args:
        wb: openpyxl Workbook
        results: List of scenario results
        run_number: Current run number
        run_summary: Summary dict for this run
    """
    sheet_name = "Run Visualizations"

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]

    # Calculate starting row for this run (each run gets ~40 rows)
    start_row = ((run_number - 1) * 45) + 1

    # Create sections
    _create_run_header(ws, start_row, run_number, run_summary)
    _create_success_metrics(ws, start_row + 3, results)
    _create_criteria_heatmap(ws, start_row + 10, results)
    _create_stage_funnel(ws, start_row + 18, results)
    _create_legend(ws, start_row + 28)

    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    for col in range(3, 20):
        ws.column_dimensions[get_column_letter(col)].width = 12


def _create_run_header(ws, row: int, run_number: int, run_summary: Dict):
    """Create the header for this run."""
    # Title
    title_cell = ws.cell(row=row, column=2, value=f"📊 RUN #{run_number} DASHBOARD")
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)

    # Timestamp
    ts_cell = ws.cell(row=row + 1, column=2, value=f"Timestamp: {run_summary['timestamp']}")
    ts_cell.font = Font(size=10, italic=True)
    ts_cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row + 1, start_column=2, end_row=row + 1, end_column=10)


def _create_success_metrics(ws, row: int, results: List[Dict]):
    """Create success rate metrics with visual indicators."""
    total = len(results)
    passed = sum(1 for r in results if r.get("success_results", {}).get("booking_confirmed", False))
    failed = total - passed
    success_rate = (passed / total * 100) if total > 0 else 0

    # Critical vs non-critical failures
    critical_failures = 0
    for r in results:
        if not r.get("success_results", {}).get("booking_confirmed", False):
            scenario = r.get("scenario", {})
            criteria_results = r.get("criteria_results", {})
            criteria_defs = scenario.get("evaluation_criteria", {})

            for crit_name, result in criteria_results.items():
                if result == "FAIL":
                    crit_def = criteria_defs.get(crit_name, {})
                    if crit_def.get("critical", False):
                        critical_failures += 1
                        break

    # Section header
    header_cell = ws.cell(row=row, column=2, value="🎯 SUCCESS METRICS")
    header_cell.font = Font(size=12, bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)

    # Metrics grid
    metrics = [
        ("Total Scenarios", total, "2F80ED"),
        ("✅ Passed", passed, COLOR_PASS),
        ("❌ Failed", failed, COLOR_FAIL),
        ("🎯 Success Rate", f"{success_rate:.1f}%", COLOR_PASS if success_rate >= 75 else "FFA726"),
        ("⚠️ Critical Failures", critical_failures, COLOR_FAIL if critical_failures > 0 else COLOR_PASS),
    ]

    for idx, (label, value, color) in enumerate(metrics):
        col = 2 + (idx % 5) * 2
        metric_row = row + 1 + (idx // 5) * 2

        # Label
        label_cell = ws.cell(row=metric_row, column=col, value=label)
        label_cell.font = Font(size=9, bold=True)
        label_cell.alignment = Alignment(horizontal="center")

        # Value
        value_cell = ws.cell(row=metric_row + 1, column=col, value=value)
        value_cell.font = Font(size=14, bold=True, color=color)
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")


def _create_criteria_heatmap(ws, row: int, results: List[Dict]):
    """Create a heatmap of criteria results across scenarios."""
    # Section header
    header_cell = ws.cell(row=row, column=2, value="🔥 CRITERIA HEATMAP")
    header_cell.font = Font(size=12, bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)

    # Collect all criteria
    all_criteria = set()
    for r in results:
        scenario = r.get("scenario", {})
        criteria = scenario.get("evaluation_criteria", {})
        all_criteria.update(criteria.keys())

    sorted_criteria = sorted(all_criteria)[:8]  # Limit to 8 for space

    # Column headers (criteria names)
    ws.cell(row=row + 1, column=2, value="Scenario").font = Font(bold=True, size=9)
    for idx, criterion in enumerate(sorted_criteria, 3):
        header = ws.cell(row=row + 1, column=idx, value=criterion.replace("_", " ")[:15])
        header.font = Font(bold=True, size=8)
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=45)
        ws.row_dimensions[row + 1].height = 60

    # Row data (scenarios)
    for scenario_idx, r in enumerate(results, row + 2):
        scenario_name = r.get("scenario", {}).get("name", "Unknown")[:20]
        ws.cell(row=scenario_idx, column=2, value=scenario_name).font = Font(size=9)

        criteria_results = r.get("criteria_results", {})
        scenario_criteria = r.get("scenario", {}).get("evaluation_criteria", {})

        for crit_idx, criterion in enumerate(sorted_criteria, 3):
            if criterion in scenario_criteria:
                result = criteria_results.get(criterion, "N/A")
                cell = ws.cell(row=scenario_idx, column=crit_idx, value=result)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=9, bold=True, color="FFFFFF")

                # Color code
                if result == "PASS":
                    cell.fill = PatternFill(start_color=COLOR_PASS, end_color=COLOR_PASS, fill_type="solid")
                elif result == "FAIL":
                    cell.fill = PatternFill(start_color=COLOR_FAIL, end_color=COLOR_FAIL, fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color=COLOR_NA, end_color=COLOR_NA, fill_type="solid")
            else:
                cell = ws.cell(row=scenario_idx, column=crit_idx, value="-")
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")


def _create_stage_funnel(ws, row: int, results: List[Dict]):
    """Create a visual funnel showing conversation stage progression."""
    from ..booking import CONVERSATION_STEPS

    # Section header
    header_cell = ws.cell(row=row, column=2, value="📈 CONVERSATION STAGE FUNNEL")
    header_cell.font = Font(size=12, bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)

    # Count scenarios at each stage
    stage_counts = {stage: 0 for stage in CONVERSATION_STEPS}
    for r in results:
        stage = r.get("success_results", {}).get("conversation_stage", "GREETING")
        if stage in stage_counts:
            stage_counts[stage] += 1

    total = len(results)

    # Display funnel
    for idx, stage in enumerate(CONVERSATION_STEPS[:6], row + 1):  # Show first 6 stages
        count = stage_counts.get(stage, 0)
        percentage = (count / total * 100) if total > 0 else 0

        # Stage name
        stage_cell = ws.cell(row=idx, column=2, value=stage.replace("_", " "))
        stage_cell.font = Font(size=9, bold=True)

        # Bar representation
        bar_length = min(int(percentage / 10), 8)  # Max 8 columns
        for bar_col in range(3, 3 + bar_length):
            bar_cell = ws.cell(row=idx, column=bar_col, value="█")
            bar_cell.font = Font(size=14, color=COLOR_PASS if stage == "BOOKING_CONFIRMED" else "2196F3")
            bar_cell.alignment = Alignment(horizontal="center")

        # Count and percentage
        stats_cell = ws.cell(row=idx, column=11, value=f"{count} ({percentage:.1f}%)")
        stats_cell.font = Font(size=9)
        stats_cell.alignment = Alignment(horizontal="right")


def _create_legend(ws, row: int):
    """Create a legend explaining the visualization."""
    # Section header
    header_cell = ws.cell(row=row, column=2, value="📖 LEGEND")
    header_cell.font = Font(size=12, bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)

    # Legend items
    legend_items = [
        ("PASS", COLOR_PASS, "Criterion met successfully"),
        ("FAIL", COLOR_FAIL, "Criterion not met"),
        ("N/A", COLOR_NA, "Criterion not applicable or not evaluated"),
        ("-", "EEEEEE", "Criterion does not apply to this scenario"),
    ]

    for idx, (label, color, description) in enumerate(legend_items, row + 1):
        # Color box
        box_cell = ws.cell(row=idx, column=2, value="  ")
        box_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        box_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Label
        label_cell = ws.cell(row=idx, column=3, value=label)
        label_cell.font = Font(bold=True, size=9)

        # Description
        desc_cell = ws.cell(row=idx, column=4, value=description)
        desc_cell.font = Font(size=9)
        ws.merge_cells(start_row=idx, start_column=4, end_row=idx, end_column=8)

    # Evaluation method note
    note_row = row + len(legend_items) + 2
    note_cell = ws.cell(row=note_row, column=2, value="💡 Evaluation Methods:")
    note_cell.font = Font(size=10, bold=True)
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=10)

    methods = [
        "• Pattern-Based: Objective criteria (name, phone, pricing) - Fast, deterministic",
        "• LLM-Based: Subjective criteria (empathy, patience, courtesy) - Contextual, accurate",
    ]

    for idx, method in enumerate(methods, note_row + 1):
        method_cell = ws.cell(row=idx, column=2, value=method)
        method_cell.font = Font(size=8, italic=True)
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=10)
