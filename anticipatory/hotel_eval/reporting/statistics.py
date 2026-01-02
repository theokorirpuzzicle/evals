"""
Historical statistics and analysis utilities.
"""

import os
import logging
from typing import Optional, Dict, List

logger = logging.getLogger("eval-runner")

DEFAULT_RESULTS_FILE = "evaluation_results.xlsx"


def get_historical_stats(excel_file: str = DEFAULT_RESULTS_FILE) -> Optional[Dict]:
    """
    Get historical statistics from the Excel file.

    Args:
        excel_file: Path to the Excel results file

    Returns:
        Dict with historical stats, or None if file doesn't exist
    """
    if not os.path.exists(excel_file):
        return None

    try:
        import openpyxl
    except ImportError:
        return None

    wb = openpyxl.load_workbook(excel_file, read_only=True)

    if "Runs Summary" not in wb.sheetnames:
        return None

    ws = wb["Runs Summary"]

    runs = []
    for row in range(2, ws.max_row + 1):
        run_data = {
            "run_number": ws.cell(row=row, column=1).value,
            "timestamp": ws.cell(row=row, column=2).value,
            "date": ws.cell(row=row, column=3).value,
            "total": ws.cell(row=row, column=4).value,
            "passed": ws.cell(row=row, column=5).value,
            "failed": ws.cell(row=row, column=6).value,
            "success_rate": ws.cell(row=row, column=7).value,
        }
        if run_data["run_number"] is not None:
            runs.append(run_data)

    if not runs:
        return None

    # Calculate overall stats
    total_runs = len(runs)
    total_scenarios = sum(r["total"] for r in runs)
    total_passed = sum(r["passed"] for r in runs)
    overall_rate = (total_passed / total_scenarios * 100) if total_scenarios > 0 else 0

    rates = [r["success_rate"] for r in runs if r["success_rate"] is not None]
    avg_rate = sum(rates) / len(rates) if rates else 0

    # Trend calculation (comparing last 3 runs to previous 3)
    trend = "stable"
    if len(rates) >= 6:
        recent = sum(rates[-3:]) / 3
        previous = sum(rates[-6:-3]) / 3
        if recent > previous + 5:
            trend = "improving"
        elif recent < previous - 5:
            trend = "declining"

    return {
        "total_runs": total_runs,
        "total_scenarios": total_scenarios,
        "total_passed": total_passed,
        "overall_success_rate": round(overall_rate, 1),
        "average_success_rate": round(avg_rate, 1),
        "best_rate": max(rates) if rates else 0,
        "worst_rate": min(rates) if rates else 0,
        "trend": trend,
        "runs": runs,
    }


def print_historical_summary(excel_file: str = DEFAULT_RESULTS_FILE):
    """
    Print a summary of historical evaluation results.

    Args:
        excel_file: Path to the Excel results file
    """
    stats = get_historical_stats(excel_file)

    if not stats:
        logger.info("No historical data found.")
        return

    logger.info(f"\n{'=' * 60}")
    logger.info("HISTORICAL EVALUATION SUMMARY")
    logger.info(f"{'=' * 60}")
    logger.info(f"Total runs: {stats['total_runs']}")
    logger.info(f"Total scenarios evaluated: {stats['total_scenarios']}")
    logger.info(f"Total passed: {stats['total_passed']}")
    logger.info(f"Overall success rate: {stats['overall_success_rate']:.1f}%")
    logger.info(f"Average success rate: {stats['average_success_rate']:.1f}%")
    logger.info(f"Best run: {stats['best_rate']:.1f}%")
    logger.info(f"Worst run: {stats['worst_rate']:.1f}%")
    logger.info(f"Trend: {stats['trend'].upper()}")

    # Show recent runs
    logger.info(f"\nRecent runs:")
    for run in stats["runs"][-5:]:
        logger.info(f"  Run #{run['run_number']} ({run['date']}): {run['success_rate']:.1f}% ({run['passed']}/{run['total']})")
