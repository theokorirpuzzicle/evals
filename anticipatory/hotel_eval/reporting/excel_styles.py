"""
Excel styling utilities for consistent formatting across reports.
"""

from typing import Optional

try:
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Color constants
COLOR_HEADER_BG = "4472C4"
COLOR_HEADER_FG = "FFFFFF"
COLOR_SUCCESS = "C6EFCE"
COLOR_WARNING = "FFEB9C"
COLOR_FAILURE = "FFC7CE"


def get_header_style():
    """Get style for header cells."""
    if not OPENPYXL_AVAILABLE:
        return None

    return {
        "font": Font(bold=True, color=COLOR_HEADER_FG),
        "fill": PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid"),
        "alignment": Alignment(horizontal="center"),
    }


def get_success_fill():
    """Get fill color for success cells."""
    if not OPENPYXL_AVAILABLE:
        return None
    return PatternFill(start_color=COLOR_SUCCESS, end_color=COLOR_SUCCESS, fill_type="solid")


def get_warning_fill():
    """Get fill color for warning cells."""
    if not OPENPYXL_AVAILABLE:
        return None
    return PatternFill(start_color=COLOR_WARNING, end_color=COLOR_WARNING, fill_type="solid")


def get_failure_fill():
    """Get fill color for failure cells."""
    if not OPENPYXL_AVAILABLE:
        return None
    return PatternFill(start_color=COLOR_FAILURE, end_color=COLOR_FAILURE, fill_type="solid")


def apply_header_style(cell):
    """Apply header styling to a cell."""
    if not OPENPYXL_AVAILABLE:
        return

    style = get_header_style()
    cell.font = style["font"]
    cell.fill = style["fill"]
    cell.alignment = style["alignment"]


def apply_rate_color(cell, rate: float):
    """
    Apply color to a rate cell based on the value.

    Args:
        cell: The openpyxl cell to style
        rate: Success rate percentage (0-100)
    """
    if not OPENPYXL_AVAILABLE:
        return

    if rate >= 80:
        cell.fill = get_success_fill()
    elif rate >= 50:
        cell.fill = get_warning_fill()
    else:
        cell.fill = get_failure_fill()


def apply_result_color(cell, is_success: bool):
    """
    Apply color to a result cell based on success/failure.

    Args:
        cell: The openpyxl cell to style
        is_success: Whether this represents a success
    """
    if not OPENPYXL_AVAILABLE:
        return

    if is_success:
        cell.fill = get_success_fill()
    else:
        cell.fill = get_failure_fill()


def apply_stage_color(cell, stage: str, step_num: int):
    """
    Apply color to a stage cell based on progress.

    Args:
        cell: The openpyxl cell to style
        stage: The conversation stage name
        step_num: The step number (1-15)
    """
    if not OPENPYXL_AVAILABLE:
        return

    if stage == "BOOKING_CONFIRMED":
        cell.fill = get_success_fill()
    elif step_num >= 10:  # Late stage
        cell.fill = get_warning_fill()
    elif step_num <= 5:  # Early failure
        cell.fill = get_failure_fill()
